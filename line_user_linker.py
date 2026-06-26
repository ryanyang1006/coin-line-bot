from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_WORKBOOK_PATH = Path(__file__).with_name("網頁資料_設計版.xlsx")


@dataclass(frozen=True)
class LineUserSettings:
    workbook_path: Path = DEFAULT_WORKBOOK_PATH
    sheet_name: str = "客戶資料"
    header_row: int = 1
    account_id_column: str = "B"
    line_name_column: str = "H"
    line_user_id_column: str = "I"


def normalize_lookup_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def parse_line_user_text(text: str) -> dict[str, str]:
    normalized = re.sub(r"\s+", " ", str(text).strip())
    match = re.search(
        r"(?:Line名稱|LINE名稱)\s*[:：]\s*(?P<line_name>.+?)\s+送評帳號ID\s*[:：]\s*(?P<review_account_id>\S+)",
        normalized,
        re.IGNORECASE,
    )
    if not match:
        raise ValueError("Expected text format: Line名稱 : xxx 送評帳號ID : xxx")
    return {
        "line_name": match.group("line_name").strip(),
        "review_account_id": match.group("review_account_id").strip(),
    }


def link_line_user(
    line_name: str,
    review_account_id: str,
    line_user_id: str,
    settings: LineUserSettings | None = None,
) -> dict[str, Any]:
    settings = settings or LineUserSettings()
    workbook_path = settings.workbook_path
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    target_id = normalize_lookup_value(review_account_id)
    if not target_id:
        raise ValueError("review_account_id is required.")
    if not str(line_name).strip():
        raise ValueError("line_name is required.")
    if not str(line_user_id).strip():
        raise ValueError("line_user_id is required.")

    workbook = load_workbook(workbook_path)
    if settings.sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {settings.sheet_name}")

    sheet = workbook[settings.sheet_name]
    first_data_row = settings.header_row + 1
    matched_row: int | None = None

    for row in range(first_data_row, sheet.max_row + 1):
        cell_value = normalize_lookup_value(sheet[f"{settings.account_id_column}{row}"].value)
        if cell_value == target_id:
            matched_row = row
            break

    if matched_row is None:
        return {
            "ok": True,
            "found": False,
            "review_account_id": target_id,
            "message": "review_account_id was not found.",
        }

    sheet[f"{settings.line_name_column}{matched_row}"] = str(line_name).strip()
    sheet[f"{settings.line_user_id_column}{matched_row}"] = str(line_user_id).strip()
    workbook.save(workbook_path)

    return {
        "ok": True,
        "found": True,
        "review_account_id": target_id,
        "matched_row": matched_row,
        "line_name_cell": f"{settings.sheet_name}!{settings.line_name_column}{matched_row}",
        "line_user_id_cell": f"{settings.sheet_name}!{settings.line_user_id_column}{matched_row}",
        "workbook_path": str(workbook_path.resolve()),
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Link LINE user information to a review account ID.")
    parser.add_argument("--line-name", required=True, help="LINE display name.")
    parser.add_argument("--review-account-id", required=True, help="Review account ID to match in 客戶資料 column B.")
    parser.add_argument("--line-user-id", required=True, help="LINE userId from the webhook event source.")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK_PATH), help="Path to 網頁資料_設計版.xlsx.")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    result = link_line_user(
        line_name=args.line_name,
        review_account_id=args.review_account_id,
        line_user_id=args.line_user_id,
        settings=LineUserSettings(workbook_path=Path(args.workbook)),
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


def normalize_line_user_payload(data: dict[str, Any]) -> dict[str, Any]:
    payload = dict(data)
    body = payload.get("body")
    if isinstance(body, str):
        body = json.loads(body)
    if isinstance(body, dict):
        payload = {**body, **{key: value for key, value in payload.items() if key != "body" and value is not None}}
    return payload


def register_line_user_routes(app: Any) -> None:
    try:
        from fastapi import Body, HTTPException
    except ImportError:
        return

    @app.post("/line-user/link")
    def link_line_user_endpoint(request: dict[str, Any] = Body(...)) -> dict[str, Any]:
        try:
            payload = normalize_line_user_payload(request)
            text = payload.get("text")
            line_user_id = payload.get("line_user_id")
            workbook_path = payload.get("workbook_path")

            if text:
                parsed = parse_line_user_text(text)
                line_name = parsed["line_name"]
                review_account_id = parsed["review_account_id"]
            else:
                line_name = payload.get("line_name")
                review_account_id = payload.get("review_account_id")
            if not line_name or not review_account_id:
                raise ValueError("Pass either text, or both line_name and review_account_id.")
            if not line_user_id:
                raise ValueError("line_user_id is required.")

            settings = LineUserSettings(
                workbook_path=Path(workbook_path) if workbook_path else DEFAULT_WORKBOOK_PATH
            )
            return link_line_user(
                line_name=line_name,
                review_account_id=review_account_id,
                line_user_id=line_user_id,
                settings=settings,
            )
        except Exception as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc


try:
    from fastapi import FastAPI

    app = FastAPI(title="LINE User Linker")
    register_line_user_routes(app)
except ImportError:
    app = None


if __name__ == "__main__":
    raise SystemExit(main())
