from __future__ import annotations

import argparse
import copy
import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_WORKBOOK_PATH = Path(os.getenv("WEB_DATA_WORKBOOK_PATH", Path(__file__).with_name("網頁資料_設計版.xlsx")))
DEFAULT_OUTPUT_DIR = Path(os.getenv("PCGS_HISTORY_OUTPUT_DIR", Path(__file__).with_name("outputs") / "pcgs_history"))
DEFAULT_FLEX_TEMPLATE_PATH = Path(os.getenv("PCGS_FLEX_TEMPLATE_PATH", Path(__file__).with_name("flex_text_sample.json")))


@dataclass(frozen=True)
class PcgsHistorySettings:
    workbook_path: Path = DEFAULT_WORKBOOK_PATH
    customer_sheet: str = "客戶資料"
    history_sheet: str = "送評紀錄"
    history_capture_sheet: str = "_pcgs_history_capture"
    customer_header_row: int = 1
    history_header_row: int = 4
    customer_name_column: str = "A"
    customer_line_user_id_column: str = "I"
    history_customer_column: str = "D"
    history_status_column: str = "I"
    history_first_column: str = "A"
    history_last_column: str = "I"
    target_status: str = "未取回"
    output_dir: Path = DEFAULT_OUTPUT_DIR
    flex_template_path: Path = DEFAULT_FLEX_TEMPLATE_PATH
    max_bubbles: int = 12
    screenshot_wait_seconds: float = 1.5
    visible_excel: bool = True


def normalize_lookup_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, date):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def format_money(value: Any) -> str:
    if value is None or value == "":
        return "$0"
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return str(value).strip()
    if amount.is_integer():
        return f"${int(amount):,}"
    return f"${amount:,.2f}"


def normalize_pcgs_history_payload(data: dict[str, Any]) -> dict[str, Any]:
    payload = dict(data)
    body = payload.get("body")
    if isinstance(body, str):
        body = json.loads(body)
    if isinstance(body, dict):
        payload = {**body, **{key: value for key, value in payload.items() if key != "body" and value is not None}}
    if "line_user_id" not in payload:
        for key in ("Line_UserId", "lineUserId", "line_userId"):
            if key in payload:
                payload["line_user_id"] = payload[key]
                break
    return payload


def parse_pcgs_history_command(text: str) -> str:
    normalized = str(text).strip()
    match = re.fullmatch(r'pcgs_history\s+["\']?(?P<line_user_id>[^"\']+)["\']?', normalized, re.IGNORECASE)
    if not match:
        raise ValueError('Expected command format: pcgs_history "Line_UserId"')
    return match.group("line_user_id").strip()


def find_customer_name(customer_ws: Any, line_user_id: str, settings: PcgsHistorySettings) -> tuple[str | None, int | None]:
    target = normalize_lookup_value(line_user_id)
    first_data_row = settings.customer_header_row + 1
    name_col = settings.customer_name_column
    id_col = settings.customer_line_user_id_column

    for row in range(first_data_row, customer_ws.max_row + 1):
        if normalize_lookup_value(customer_ws[f"{id_col}{row}"].value) == target:
            name = normalize_lookup_value(customer_ws[f"{name_col}{row}"].value)
            return (name or None), row
    return None, None


def build_text_message(text: str) -> dict[str, str]:
    return {
        "type": "text",
        "text": text,
    }


def collect_matching_history_rows(history_ws: Any, customer_name: str, settings: PcgsHistorySettings) -> list[dict[str, Any]]:
    first_data_row = settings.history_header_row + 1
    rows: list[dict[str, Any]] = []
    for row in range(first_data_row, history_ws.max_row + 1):
        row_customer = normalize_lookup_value(history_ws[f"{settings.history_customer_column}{row}"].value)
        row_status = normalize_lookup_value(history_ws[f"{settings.history_status_column}{row}"].value)
        if row_customer == customer_name and row_status == settings.target_status:
            rows.append(
                {
                    "row": row,
                    "serial_number": format_cell_value(history_ws[f"A{row}"].value),
                    "order_number": format_cell_value(history_ws[f"B{row}"].value),
                    "submitted_date": format_cell_value(history_ws[f"C{row}"].value),
                    "customer_name": row_customer,
                    "company": format_cell_value(history_ws[f"E{row}"].value),
                    "service_level": format_cell_value(history_ws[f"F{row}"].value),
                    "quantity": format_cell_value(history_ws[f"G{row}"].value),
                    "fee": format_money(history_ws[f"H{row}"].value),
                    "status": row_status,
                }
            )
    return rows


def load_flex_template(settings: PcgsHistorySettings) -> dict[str, Any]:
    return json.loads(settings.flex_template_path.read_text(encoding="utf-8-sig"))


def set_labeled_text(bubble: dict[str, Any], label: str, value: str) -> None:
    def walk(node: Any) -> bool:
        if isinstance(node, dict):
            contents = node.get("contents")
            if node.get("type") == "box" and isinstance(contents, list) and len(contents) >= 2:
                first = contents[0]
                second = contents[1]
                if isinstance(first, dict) and isinstance(second, dict) and first.get("text") == label:
                    second["text"] = value
                    return True
            return any(walk(value) for value in node.values())
        if isinstance(node, list):
            return any(walk(item) for item in node)
        return False

    walk(bubble)


def set_metric_text(bubble: dict[str, Any], label: str, value: str) -> None:
    def walk(node: Any) -> bool:
        if isinstance(node, dict):
            contents = node.get("contents")
            if node.get("type") == "box" and isinstance(contents, list):
                for index, item in enumerate(contents[:-1]):
                    if isinstance(item, dict) and item.get("text") == label:
                        next_item = contents[index + 1]
                        if isinstance(next_item, dict):
                            next_item["text"] = value
                            return True
            return any(walk(value) for value in node.values())
        if isinstance(node, list):
            return any(walk(item) for item in node)
        return False

    walk(bubble)


def build_history_bubble(template_bubble: dict[str, Any], row: dict[str, Any]) -> dict[str, Any]:
    bubble = copy.deepcopy(template_bubble)
    card_box = bubble["body"]["contents"][2]
    card_box["contents"][0]["text"] = row["order_number"] or "未填送評單號"
    set_labeled_text(bubble, "送評日期", row["submitted_date"])
    set_labeled_text(bubble, "委託人", row["customer_name"])
    set_labeled_text(bubble, "送評公司", row["company"])
    set_labeled_text(bubble, "等級", row["service_level"])
    set_labeled_text(bubble, "收款狀態", row["status"])
    set_metric_text(bubble, "數量", row["quantity"])
    set_metric_text(bubble, "費用", row["fee"])
    bubble["action"] = {
        "type": "message",
        "text": f"送評單號:{row['order_number']}",
    }
    return bubble


def build_flex_message(rows: list[dict[str, Any]], settings: PcgsHistorySettings) -> dict[str, Any]:
    template = load_flex_template(settings)
    if not template.get("contents"):
        raise ValueError("flex_text_sample.json must contain at least one bubble.")
    template_bubble = template["contents"][0]
    bubbles = [build_history_bubble(template_bubble, row) for row in rows[: settings.max_bubbles]]
    carousel = copy.deepcopy(template)
    carousel["contents"] = bubbles
    return {
        "type": "flex",
        "altText": "送評紀錄",
        "contents": carousel,
    }


def export_pcgs_history(line_user_id: str, settings: PcgsHistorySettings | None = None) -> dict[str, Any]:
    settings = settings or PcgsHistorySettings()
    if not settings.workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {settings.workbook_path}")

    target_line_user_id = normalize_lookup_value(line_user_id)
    if not target_line_user_id:
        raise ValueError("line_user_id is required.")

    workbook = load_workbook(settings.workbook_path, read_only=True, data_only=True)
    try:
        customer_ws = workbook[settings.customer_sheet]
        history_ws = workbook[settings.history_sheet]
        customer_name, customer_row = find_customer_name(customer_ws, target_line_user_id, settings)

        if not customer_name:
            return {
                "ok": True,
                "found": False,
                "message": "line user not found",
                "line_user_id": target_line_user_id,
                "messages": [build_text_message("查無綁定的 LINE 使用者資料。")],
            }

        history_rows = collect_matching_history_rows(history_ws, customer_name, settings)
        if not history_rows:
            return {
                "ok": True,
                "found": True,
                "has_history": False,
                "message": "no history found",
                "line_user_id": target_line_user_id,
                "customer_name": customer_name,
                "customer_row": customer_row,
                "messages": [build_text_message("目前沒有符合條件的送評紀錄。")],
            }

        flex_message = build_flex_message(history_rows, settings)
        return {
            "ok": True,
            "found": True,
            "has_history": True,
            "message": "history found",
            "line_user_id": target_line_user_id,
            "customer_name": customer_name,
            "customer_row": customer_row,
            "history_rows": len(history_rows),
            "returned_bubbles": len(flex_message["contents"]["contents"]),
            "truncated": len(history_rows) > settings.max_bubbles,
            "flex_message": flex_message,
            "messages": [flex_message],
            "workbook_path": str(settings.workbook_path.resolve()),
        }
    finally:
        workbook.close()


def register_pcgs_history_routes(app: Any) -> None:
    try:
        from fastapi import Body, HTTPException
    except ImportError:
        return

    @app.post("/pcgs-history")
    def pcgs_history_endpoint(request: dict[str, Any] = Body(...)) -> dict[str, Any]:
        try:
            payload = normalize_pcgs_history_payload(request)
            command = payload.get("command") or payload.get("text")
            line_user_id = payload.get("line_user_id")
            if command and not line_user_id:
                line_user_id = parse_pcgs_history_command(command)
            workbook_path = payload.get("workbook_path")
            settings = PcgsHistorySettings(
                workbook_path=Path(workbook_path) if workbook_path else DEFAULT_WORKBOOK_PATH
            )
            return export_pcgs_history(line_user_id=line_user_id, settings=settings)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Export unreturned PCGS history for a LINE user.")
    parser.add_argument("--line-user-id", help="LINE userId to look up in 客戶資料 column I.")
    parser.add_argument("--command", help='LINE-style command, e.g. pcgs_history "Line_UserId".')
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK_PATH), help="Path to 網頁資料_設計版.xlsx.")
    return parser


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    args = build_parser().parse_args()
    line_user_id = args.line_user_id or (parse_pcgs_history_command(args.command) if args.command else None)
    result = export_pcgs_history(
        line_user_id=line_user_id,
        settings=PcgsHistorySettings(workbook_path=Path(args.workbook)),
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


try:
    from fastapi import FastAPI

    app = FastAPI(title="PCGS History Bot")
    register_pcgs_history_routes(app)
except ImportError:
    app = None


if __name__ == "__main__":
    raise SystemExit(main())
